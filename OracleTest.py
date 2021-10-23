import openpyxl
#
path="D:\\New.xlsx"
workbook=openpyxl.load_workbook(path)
# sheet_nm=workbook.sheetnames
# print(sheet_nm)
# #sheet_nm=workbook.sheetnames
# #print(sheet_nm)
# # sheet_4=workbook['Sheet4']
# # data=sheet_4['B3'].value
# # print(data)
# #
# # data1=workbook['Sheet4']['A2'].value
# # print(data1)
# # print(sheet_4.cell(7,4).value)
#
# # sheet_nm1=workbook['Sheet5']
# # val1=sheet_nm1[21][0].value
# # print(val1)
# # val2=sheet_nm1.cell(11,11).value
# # print(val2)
# #
# # get_sheet=workbook.get_sheet_by_name("Sheet6").cell(row=9,column=1).value
# # print(get_sheet)
#
#
sheet=workbook['Corona list']
rows=sheet.max_row
coulmns=sheet.max_column

print(rows)
print(coulmns)

for r in range(2,rows+1):
    for c in range(1,coulmns+1):
        check=sheet.cell(row=r,column=c).value
        print(check)
#
# column_list=sheet['C12'].value
# print(column_list)
# print(workbook.get_sheet_by_name('Corona list').cell(row=12,column=2).value)
#
# sheet.cell(row=20,column=1,value="333")
# sheet.cell(row=20,column=2,value="Hydrabad")
# sheet.cell(row=20,column=3,value="Gachibowli")
#
# workbook.save("D:\\New.xlsx")
#
# root="D:\\New.xlsx"
# wb=openpyxl.load_workbook(root)
# sn=wb.sheetnames
# print(sn)
#
# sheet=wb.get_sheet_by_name("funds").cell(row=1,column=1).value
# print(sheet)
#
# workb=wb["funds"]['C6'].value
# print(workb)



