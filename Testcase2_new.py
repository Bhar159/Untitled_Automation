from selenium import webdriver
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
import selenium
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl



chromedriver = 'D:\Selenium\chromedriver_win32/chromedriver'
driver = webdriver.Chrome(chromedriver)
driver.get("https://anuntaqa.mydesktopnow.com/customerOnboard")
driver.maximize_window()
driver.find_element_by_xpath("//div[@class='p-4 ng-star-inserted']/form/div[1]/div/input").send_keys('QA')
driver.find_element_by_xpath("//div[@class='p-4 ng-star-inserted']/form/div[2]/div/input").send_keys("pass@123")
driver.find_element_by_xpath("//div[@class='p-4 ng-star-inserted']/form/div[2]/div/input").forward()
javascript=driver.find_element_by_xpath("//button/span[contains(text(),'Login')]")
driver.execute_script("arguments[0].click();", javascript)
#driver.find_element_by_xpath("//button/span[contains(text(),'Login')]").click()
time.sleep(5)
driver.find_element_by_xpath("//span[contains(text(),'Subscriptions')]").click()
time.sleep(2)
driver.find_element_by_xpath("//div[@class='mat-tab-list']/div/div[2]").click()
time.sleep(2)
time.sleep(2)
driver.find_element_by_xpath("//div[@class='ng-star-inserted']/div[2]/div/div[2]/div/div/div[1]/div/div[1]/div[2]/input").click()
time.sleep(2)

azureADtext=driver.find_element_by_xpath("//div[@class='ng-star-inserted']/div[3]/div/div[2]/div/div/div/div[contains(text(),'Azure AD')]")


if (azureADtext.is_displayed):
    print("second one started")
    # print("working fine")


else:
    print("started")
    existingAccount = driver.find_element_by_xpath(
        "//div[@class='row py-4 ng-star-inserted']/div[2]/div/div/div//span[contains(text(),' You are currently logged in as')]")
    # driver.find_element_by_xpath().click()
    # time.sleep(5)
    # self.driver.find_element_by_xpath(self.elogoutFirstoption_xpath).click()
    # time.sleep(5)
    # self.drAddSubscriptionTab()
    # time.sleep(5)
    # self.sharedSubCheckBox()



# driver.find_element(By.XPATH,"//span[contains(text(),'Customers')]").click()
# waitUp=WebDriverWait(driver,10)
# #driver.find_element_by_xpath("//div[contains(text(),'Subscriptions')]")
# textSub=waitUp.until(ec.text_to_be_present_in_element((By.XPATH,"//div[contains(text(),'Subscriptions')]"),'Subscriptions'))
# print(textSub)
#
#
#
# # key=driver.find_element_by_xpath("//div[@class='p-4 ng-star-inserted']/form/div[1]/div/input")
# # actions=ActionChains(driver)
# # key.send_keys("kolo")
# # # actions.click(key).perform()
# # # actions.click(key).send_keys("Hello").perform()
# # # time.sleep(5)
# # # actions.send_keys_to_element(key,'hola')
# # actions.key_down(Keys.CONTROL).send_keys("A").key_up(Keys.CONTROL).perform()
# # actions.key_down(Keys.CONTROL).send_keys("C").key_up(Keys.CONTROL).perform()
# # password=driver.find_element_by_xpath("//div[@class='p-4 ng-star-inserted']/form/div[2]/div/input")
# actions.click(password).key_down(Keys.CONTROL).send_keys("V").perform()
# driver.find_element_by_xpath("//button/span[contains(text(),'Login')]").click()
# time.sleep(5)
# recenttext=driver.find_element_by_xpath("//div[@id='recent']/div[3]/div")
# print("test 1")


#
# driver.execute_script("arguments[0].scrollIntoView()",recenttext)
# time.sleep(5)
# driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
# time.sleep(10)
# driver.execute_script("window.scrollTo(0, -document.body.scrollHeight)")
# driver.find_element(By.XPATH,"//span[contains(text(),'Customers')]").click()
# driver.execute_script("window.open('https://www.facebook.com/','new window')")
# time.sleep(5)
#driver.switch_to.window(driver.window_handles[0])



#driver.execute_script("arguments[0].scrollIntoView();",frame)

#frame=driver.find_element_by_xpath("//div[@class='ng-star-inserted']/div[2]/div/div[2]")
#driver.execute_script("document.querySelector('//div[@id=\"recent\"]').scrollTop=500");
#driver.execute_script("window.scrollTo(0,500);")
#time.sleep(10)
# print("Hello")
# driver.find_element_by_xpath("//span[contains(text(),'Customers')]").click()
# time.sleep(5)
# tablelist=driver.find_element_by_xpath("//div[@class='col']/div[4]/div/div/table").text
# #print(tablelist)
# colour=driver.find_element_by_xpath("//div[@class='col']/div[4]/div/div/table/tbody/tr[1]/td[2]/div[1]")
# backgroundcolour=colour.value_of_css_property('background-color')
# print(backgroundcolour)
# colour="rgba(255, 0, 0, 0.25)"
#
# actionChain=ActionChains(driver)
# actionChain.move_to_element(driver.find_element_by_xpath("//div[@class='col']/div[4]/div/div/table/tbody/tr[1]/td[2]")).perform()
# keyvalue=driver.find_element_by_xpath("//div[@class='mat-tooltip ng-trigger ng-trigger-state']").is_displayed()
# text=driver.find_element_by_xpath("//div[@class='mat-tooltip ng-trigger ng-trigger-state']").text
# print(keyvalue)
#
#
# if keyvalue==True:
#     print("successs")
#     assert text in 'Subscription preparation failed for the customer','Validation failed'
#     print("Validation Pass")
# else:
#     print("Failed")
#
#
#
# # countcustomerlist=driver.find_elements_by_xpath("//div[@class='col']/div[4]/div/div/table/tbody/tr/td[2]")
# # for x in countcustomerlist:
#     print(x.text)
#
# if backgroundcolour != colour:
#     print("First step")
#     driver.find_element_by_xpath("//div[@class='col']/div[4]/div/div/table/tbody/tr[1]/td[2]").click()
#     print("second step")
# else:
#     print("working man")
#     driver.find_element_by_xpath("//div[@class='col']/div[4]/div/div/table/tbody/tr[2]/td[2]").click()
#     print("It's fine")




# driver.find_element_by_xpath("//span[contains(text(),'VM Images')]").click()
# time.sleep(5)
# driver.find_element_by_xpath("//div[@class='col-12']/div/div/mat-tab-group/mat-tab-header/div[2]/div/div/div[2]/div").click()
# time.sleep(5)
# driver.find_element_by_xpath("//div[@class='col-sm-10']/input").send_keys("NewImage")
# driver.find_element_by_xpath("//mat-tab-body/div[1]/div[2]/div[1]/div[1]/div[2]/form[1]/div[2]/div[1]/textarea[1]").send_keys("NewImage")
# driver.find_element_by_xpath("//div[@class='col-12 pl-0 pr-0']/div/div[1]/div[1]").click()
#
# multiselect = Select(driver.find_element_by_xpath("//mat-tab-body/div/div[2]/div[2]/div[2]/form/div[1]/div/select"))
# textSelect=multiselect.options
# print(textSelect)
# for x in textSelect:
#     print(x.text)
#



# path="C:\\Users\\bhara\\PycharmProjects\\untitled\\BDD Automation\\TestData\\New.xlsx"
# wk=openpyxl.load_workbook(path)
# sheetName=wk.sheetnames
# print(sheetName)
# sheet1=wk["Corona list"]
# x=sheet1.max_row
# y=sheet1.max_column
# print(x)
# print(y)
#
#
# for u in range(1,x+1):
#     for n in range(1,y+1):
#         sheet01=sheet1.cell(row=u,column=n).value
#         print(sheet01)







